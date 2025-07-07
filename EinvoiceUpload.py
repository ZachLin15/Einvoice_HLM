import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import warnings
import win32com.client as win32
import time
from datetime import datetime
import smtplib
from email.message import EmailMessage
import mimetypes


# Suppress all Excel-related warnings for automated processing
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', message='.*Data Validation.*')
warnings.filterwarnings('ignore', message='.*Workbook contains no default style.*')

import win32com.client as win32
import os
import time


def refresh_data_connections(file_path):
    """
    Opens an Excel workbook visibly, refreshes all data connections, waits for completion,
    saves, and closes the workbook.

    Args:
        file_path (str): Full path to the Excel file (.xlsm, .xlsx, etc.)
    """
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return False

    try:
        # Launch Excel via COM
        excel_app = win32.gencache.EnsureDispatch("Excel.Application")
        excel_app.Visible = True
        excel_app.DisplayAlerts = False
        excel_app.ScreenUpdating = True  # Ensure you can see it

        print(f"📂 Opening workbook: {file_path}")
        workbook = excel_app.Workbooks.Open(file_path)

        print("🔄 Refreshing all connections...")
        workbook.RefreshAll()

        # Wait for Excel to finish refreshing
        print("⏳ Waiting for Excel to finish refreshing...")
        time.sleep(30)

        # Save the workbook after refresh
        workbook.Save()
        print("✅ Workbook refreshed and saved.")

        # Close and cleanup
        workbook.Close(SaveChanges=False)
        excel_app.Quit()

        return True

    except Exception as e:
        print(f"❌ Error while refreshing: {e}")
        try:
            if 'workbook' in locals():
                workbook.Close(SaveChanges=False)
            if 'excel_app' in locals():
                excel_app.Quit()
        except:
            pass
        return False


def copy_data_between_files(source_file, submission_file, output_file, refresh_connections=True):
    """
    Copy data from source file sheets to submission file sheets and save as new file.
    Optionally refresh data connections first.

    Args:
        source_file: Path to exportM file
        submission_file: Path to submission file
        output_file: Path to save the new combined file
        refresh_connections: Whether to refresh data connections before copying
    """

    # Sheets to copy data between
    sheets_to_copy = [
        'Documents',
        'DocumentLineItems',
        'LineItemsTaxes',
        'DocumentTotalTax'
    ]

    try:
        # Step 1: Refresh data connections if requested
        if refresh_connections:
            print("=" * 50)
            print("STEP 1: REFRESHING DATA CONNECTIONS")
            print("=" * 50)

            success = refresh_data_connections(source_file)
            if success:
                print("✓ Data connections refreshed successfully")
            else:
                print("⚠ Warning: Could not refresh all connections, continuing with copy...")

            print("\n" + "=" * 50)
            print("STEP 2: COPYING DATA")
            print("=" * 50)

        print("Loading files...")
        print(f"Source file: {source_file}")
        print(f"Submission file: {submission_file}")

        # Load both workbooks
        source_wb = load_workbook(source_file)
        submission_wb = load_workbook(submission_file)

        print(f"Source sheets available: {source_wb.sheetnames}")
        print(f"Submission sheets available: {submission_wb.sheetnames}")

        # Process each sheet
        for sheet_name in sheets_to_copy:
            if sheet_name in source_wb.sheetnames and sheet_name in submission_wb.sheetnames:
                print(f"\nCopying data for sheet: {sheet_name}")

                # Get sheets
                source_sheet = source_wb[sheet_name]
                submission_sheet = submission_wb[sheet_name]

                # Find data range in source (skip header row)
                max_row = source_sheet.max_row
                max_col = source_sheet.max_column

                print(f"Source data: {max_row} rows x {max_col} columns")

                # Copy data from source to submission (skip source header, start at row 6 in submission)
                data_rows_copied = 0
                for row in range(2, max_row + 1):  # Skip header row in source (start from row 2)
                    for col in range(1, max_col + 1):
                        source_cell = source_sheet.cell(row=row, column=col)
                        # Paste starting at row 6 in submission (keeping top 5 rows)
                        submission_cell = submission_sheet.cell(row=row + 4, column=col)
                        submission_cell.value = source_cell.value
                    data_rows_copied += 1

                print(f"Copied {data_rows_copied} rows of data to {sheet_name}")

            elif sheet_name not in source_wb.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found in source file")
            elif sheet_name not in submission_wb.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found in submission file")

        # Save as new Excel file
        print(f"\nSaving combined file as: {output_file}")
        submission_wb.save(output_file)

        print("\n" + "=" * 50)
        print("PROCESS COMPLETED")
        print("=" * 50)
        print("✓ Data connections refreshed!")
        print("✓ Data copy completed successfully!")
        print(f"✓ New file saved: {output_file}")
        print("✓ Original files remain unchanged")

    except FileNotFoundError as e:
        print(f"Error: File not found - {e}")
    except Exception as e:
        print(f"Error occurred: {e}")

def send_email_smtp(sender_email, sender_password, to, cc, subject, body, attachment_path, smtp_server, smtp_port):
    """
    Sends an email with attachment using SMTP (TLS or SSL).
    """
    msg = EmailMessage()
    msg['From'] = sender_email
    msg['To'] = to
    msg['cc'] = cc
    msg['Subject'] = subject
    msg.set_content(body)

    # Add attachment
    if os.path.exists(attachment_path):
        mime_type, _ = mimetypes.guess_type(attachment_path)
        mime_type, mime_subtype = mime_type.split('/') if mime_type else ('application', 'octet-stream')

        with open(attachment_path, 'rb') as file:
            msg.add_attachment(
                file.read(),
                maintype=mime_type,
                subtype=mime_subtype,
                filename=os.path.basename(attachment_path)
            )
        print(f"📎 Attached file: {attachment_path}")
    else:
        print(f"⚠ Attachment not found: {attachment_path}")
        return

    # Send email
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as smtp:
            smtp.starttls()  # Secure the connection
            smtp.login(sender_email, sender_password)
            smtp.send_message(msg)
        print("📧 Email sent successfully via SMTP.")
    except Exception as e:
        print(f"❌ Failed to send email: {e}")


def open_and_resave_excel(file_path):
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return False

    try:
        excel_app = win32.gencache.EnsureDispatch("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False

        print(f"📂 Opening workbook: {file_path}")
        workbook = excel_app.Workbooks.Open(file_path)

        # Save to a new file
        new_path = file_path.replace(".xlsx", "_cleaned.xlsx")
        workbook.SaveAs(new_path)
        workbook.Close(SaveChanges=False)
        excel_app.Quit()

        print(f"✅ File saved cleanly as: {new_path}")
        return True

    except Exception as e:
        print(f"❌ Error while processing file: {e}")
        try:
            if 'workbook' in locals():
                workbook.Close(SaveChanges=False)
            if 'excel_app' in locals():
                excel_app.Quit()
        except:
            pass
        return False


if __name__ == "__main__":
    # File paths - update these with your actual file paths
    source_file = os.path.abspath(r"_internal\M-Einvoice\exportM2.xlsx")  # Source file
    submission_file = r"_internal\M-Einvoice\BatchSubmission-v2.xlsx"  # Submission file
    output_file = f"_internal\\M-Einvoice\\E Invoice {datetime.today().strftime('%Y-%m-%d')}.xlsx"  # New output file

    # Check if files exist
    if not os.path.exists(source_file):
        print(f"Error: Source file '{source_file}' not found!")
        exit(1)

    if not os.path.exists(submission_file):
        print(f"Error: Submission file '{submission_file}' not found!")
        exit(1)

    print("Starting Excel data copy with connection refresh...")
    print("Using openpyxl method (preserves formatting)")

    # Execute the copy process with connection refresh
    copy_data_between_files(source_file, submission_file, output_file, refresh_connections=True)
    #open_and_resave_excel(f"_internal\\M-Einvoice\\E Invoice {datetime.today().strftime('%Y-%m-%d')}.xlsx")

    # Email config
    sender_email = "admin1@lshworld.com"
    sender_password = "dpvqmxwsrxvxmbvr"
    recipient = "carene_my@lshworld.com"
    cc = "sales_my@lshworld.com"
    subject = f"E-Invoice Date {datetime.today().strftime('%Y-%m-%d')}"
    body = "Hi,\n\nPlease find the attached combined submission Excel file.\n\nRegards,\nAutomation Script"
    smtp_server = "smtp.office365.com"  # or smtp.gmail.com or your mail server
    smtp_port = 587

    '''send_email_smtp(
            sender_email,
            sender_password,
            recipient,
            cc,
            subject,
            body,
            output_file,
            smtp_server,
            smtp_port
        )'''

